from colorama import Back, Fore, init
from openpyxl import load_workbook

ROWS, COLS = 10,10
WORKBOOK ,WORKSHEET ='aleatorios.xlsx','aleatorio'

wb = load_workbook (WORKBOOK)
ws = wb[WORKSHEET]

init()

for r in range (1, ROWS +1):
    print(
        Fore.BLUE if r % 2 else Fore.YELLOW,
        " | ".join(
            [
                '{:5d}'.format(
                    ws.cell(row=r, column=c).value
                )
                for c in range (1, COLS + 1)
            ]
        ), 
        Fore.RESET, sep=""
    )